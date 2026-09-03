"""
access.py — translates the authenticated user's Entra ID role(s) into what they
can see.

Static Web Apps injects the authenticated principal as a base64 JSON blob in the
`x-ms-client-principal` header. We read the user's roles from it. No secrets, no
token validation needed here — SWA authenticates the request before it reaches
the Function.

==============================================================================
HOW ACCESS IS DEFINED  (this is the part you edit to add/adjust groups)
==============================================================================
Every group is one entry in GROUPS below. A group declares:
  - reports:  the set of report keys it may open
              (or the sentinel ALL_REPORTS to mean "every report")
  - clinic_scoped: if True, users of this group only see THEIR OWN clinic's
              rows/tabs within a report. Their clinic(s) come from
              config/clinic_access.csv (email -> clinic) and/or legacy
              apollo_cd_<clinic> roles.

Invite users to the exact lowercase role name (the dict key), e.g. apollo_et.

CLINIC STAFF (the apollo_clinic group):
  - Invite the user with role `apollo_clinic` (one role for all 20 clinics).
  - Add a row to config/clinic_access.csv in the apollo-reports container:
        email,clinic
        jsmith@apollobehavior.com,Buford
    Clinic names must match the canonical names used in the reports
    (same spelling as clinics.csv, e.g. "Warner Robins - Byrd").
  - What they see is defined tab-by-tab in CLINIC_TAB_RULES below.
    DEFAULT DENY: any tab not listed there is hidden from clinic users,
    so a new tab added to a report is invisible to clinics until you
    explicitly grant it here.
  - AR, Daily Revenue, and Direct Labor Margin are not in the clinic
    report set, so they never appear for clinic users at all.
==============================================================================
"""

import json
import base64
import csv
import io

# ---- report catalog --------------------------------------------------------
REPORT_CATALOG = {
    "ar":          {"title": "Accounts Receivable", "scope": "all", "folder": "outputs/ar"},
    "rev":         {"title": "Daily Revenue",       "scope": "all", "folder": "outputs/revenue"},
    "supervision": {"title": "Supervision Ratio",   "scope": "all", "folder": "outputs/supervision"},
    "bcba":        {"title": "BCBA Billing",        "scope": "all", "folder": "outputs/bcba-billing"},
    "direct-labor-margin": {"title": "Direct Labor Margin", "scope": "all", "folder": "outputs/direct-labor-margin"},
    "parent-training":     {"title": "Parent Training",     "scope": "all", "folder": "outputs/parent-training"},
    "group-rate":          {"title": "RBT Group Hour Rate", "scope": "all", "folder": "outputs/group-rate"},
}

CONTAINER = "apollo-reports"

# Sentinel meaning "every report in the catalog".
ALL_REPORTS = "__ALL__"

# Reports a clinic-scoped user may open. AR / rev / DLM intentionally absent.
CLINIC_REPORTS = {"supervision", "bcba", "parent-training", "group-rate"}

# ---- GROUP DEFINITIONS (edit here to add/adjust access) --------------------
GROUPS = {
    # Executive Team: sees everything. The default superuser group.
    "apollo_et": {
        "reports": ALL_REPORTS,
        "clinic_scoped": False,
    },
    # Revenue Cycle Management: only these reports today.
    # To grant more later, add keys to this set (e.g. add "supervision").
    "apollo_rcm": {
        "reports": {"ar", "rev"},
        "clinic_scoped": False,
    },
      # Clinical Excellence: the four clinical reports, full data, no filtering.
    "apollo_clinical": {
        "reports": CLINIC_REPORTS,
        "clinic_scoped": False,
    },
    # Clinic staff: one role for all clinics. WHICH clinic a user belongs to
    # comes from config/clinic_access.csv (see header comment).
    "apollo_clinic": {
        "reports": CLINIC_REPORTS,
        "clinic_scoped": True,
    },
}

# ---- Legacy clinic roles (apollo_cd_<clinic>) -------------------------------
# Still honored: a user invited as apollo_cd_<clinic> gets the same clinic view
# as apollo_clinic, with <clinic> taken from the role name.
CD_PREFIX = "apollo_cd_"
CLINIC_VIEW = {
    "reports": CLINIC_REPORTS,
    "clinic_scoped": True,
}

# ---- Clinic mapping file -----------------------------------------------------
# CSV in the apollo-reports container: header `email,clinic`, one row per
# (user, clinic). A user listed twice gets access to both clinics.
CLINIC_ACCESS_BLOB = "config/clinic_access.csv"


# =============================================================================
# CLINIC TAB RULES — what a clinic-scoped user sees inside each report.
#
#   ("all",)               whole tab, every clinic's rows
#   ("filter", "<Column>") only rows where <Column> equals one of the user's
#                          clinics (matched case-insensitively). Rows above the
#                          header (titles) are kept. 'Unmapped' and totals rows
#                          disappear because they don't match any clinic.
#   not listed             tab is removed entirely (DEFAULT DENY)
#
# "bcba" is special: the workbook has one tab per clinic, so instead of row
# filtering we keep only the tab(s) named for the user's clinic(s). The
# All BCBAs Summary tab (all clinics) is intentionally not shown.
# =============================================================================
CLINIC_TAB_RULES = {
    "supervision": {
        "Supervision Ratio":       ("all",),
        "Supervision by Client":   ("filter", "Center"),
        "Methodology":             ("all",),
        # "Clinic Mapping" intentionally absent -> hidden
    },
    "parent-training": {
        "Parent Training":           ("all",),
        "Parent Training by Client": ("filter", "Center"),
        "Methodology":               ("all",),
        # "Clinic Mapping" intentionally absent -> hidden
    },
    "group-rate": {
        "Apollo Weekly":        ("all",),
        "Clinic Rate by Week":  ("all",),
        "RBT Rate by Week":     ("filter", "Clinic"),
        "Clinic Group Hrs":     ("all",),
        "Clinic Total Hrs":     ("all",),
        "RBT Group Hrs":        ("filter", "Clinic"),
        "RBT Total Hrs":        ("filter", "Clinic"),
        "How This Works":       ("all",),
    },
    "bcba": "own_tab",
}

# BCBA workbooks generated before Sep 2026 used short codes as tab names.
# Newer workbooks use full clinic names. Accept both when picking the user's tab.
BCBA_TAB_CODE_TO_CLINIC = {
    "AC": "Acworth", "AT": "Athens", "AU": "Austell",
    "CA": "Canton", "CT": "Canton",
    "CB": "Columbus", "CO": "Columbus",
    "CM": "Cumming", "DC": "Dacula", "DL": "Duluth", "DS": "Dallas",
    "EC": "East Cobb", "HM": "Hamilton Mill", "JC": "Johns Creek",
    "LV": "Lawrenceville", "NA": "North Alpharetta", "OW": "Oakwood",
    "SA": "South Alpharetta", "SB": "Stockbridge", "SH": "Sugar Hill",
    "WR1": "Warner Robins - Margie", "WR2": "Warner Robins - Byrd",
    # legacy long spellings from older generator versions
    "WARNER ROBINS 1- MARGIE": "Warner Robins - Margie",
    "WARNER ROBINS 2- BYRD":   "Warner Robins - Byrd",
}


def _norm(s) -> str:
    return str(s).strip().casefold()


def _principal_from_header(header_val: str) -> dict:
    if not header_val:
        return {}
    try:
        return json.loads(base64.b64decode(header_val).decode("utf-8"))
    except Exception:
        return {}


def get_user(req_headers) -> dict:
    """
    Returns:
      name, roles[], clinics[] (from apollo_cd_<clinic> roles only; csv-based
      clinics are resolved separately via resolve_user_clinics),
      is_et (bool), reports (resolved set of allowed keys or ALL_REPORTS),
      clinic_scoped (bool)
    """
    principal = _principal_from_header(req_headers.get("x-ms-client-principal"))
    name = principal.get("userDetails", "")
    roles = set(principal.get("userRoles", []))
    for c in principal.get("claims", []):
        if c.get("typ", "").endswith("groups") or c.get("typ") == "groups":
            roles.add(c.get("val"))

    is_et = "apollo_et" in roles
    clinics = sorted(r[len(CD_PREFIX):] for r in roles if r.startswith(CD_PREFIX))

    # Resolve the union of everything this user's groups allow.
    allowed = set()
    clinic_scoped = False
    grants_all = False

    for role in roles:
        g = GROUPS.get(role)
        if g:
            if g["reports"] == ALL_REPORTS:
                grants_all = True
            else:
                allowed |= set(g["reports"])
            clinic_scoped = clinic_scoped or g["clinic_scoped"]

    if clinics:  # any apollo_cd_<clinic> role grants the clinic view
        allowed |= set(CLINIC_VIEW["reports"])
        clinic_scoped = clinic_scoped or CLINIC_VIEW["clinic_scoped"]

    reports = ALL_REPORTS if grants_all else allowed

    # ET (or any all-reports group) is never clinic-scoped (sees full data).
    if grants_all:
        clinic_scoped = False

    return {
        "name": name,
        "roles": sorted(roles),
        "is_et": is_et,
        "clinics": clinics,
        "reports": reports,
        "clinic_scoped": clinic_scoped,
    }


def load_clinic_access(container_client) -> dict:
    """
    Read config/clinic_access.csv from the apollo-reports container.
    Returns {email(lowercased): [clinic, ...]}. Missing/broken file -> {}.
    """
    try:
        raw = container_client.get_blob_client(CLINIC_ACCESS_BLOB).download_blob().readall()
        text = raw.decode("utf-8-sig")  # tolerate Excel's BOM
        out = {}
        for row in csv.DictReader(io.StringIO(text)):
            email = _norm(row.get("email", ""))
            clinic = str(row.get("clinic", "")).strip()
            if email and clinic:
                out.setdefault(email, []).append(clinic)
        return out
    except Exception:
        return {}


def resolve_user_clinics(user: dict, container_client) -> list:
    """
    Union of clinics from legacy apollo_cd_<clinic> roles and from
    clinic_access.csv (matched by the signed-in email, case-insensitive).
    """
    clinics = list(user.get("clinics") or [])
    mapping = load_clinic_access(container_client)
    for c in mapping.get(_norm(user.get("name", "")), []):
        if _norm(c) not in {_norm(x) for x in clinics}:
            clinics.append(c)
    return clinics


def _may_open(user: dict, key: str) -> bool:
    if user["reports"] == ALL_REPORTS:
        return True
    return key in user["reports"]


def visible_reports(user: dict) -> list:
    out = []
    for key, cfg in REPORT_CATALOG.items():
        if _may_open(user, key):
            out.append({"key": key, "title": cfg["title"], "clinic": "all"})
    return out


def can_access(user: dict, report_key: str, clinic: str = "all") -> bool:
    if report_key not in REPORT_CATALOG:
        return False
    return _may_open(user, report_key)


def folder_prefix(report_key: str, clinic: str = "all") -> str:
    cfg = REPORT_CATALOG[report_key]
    base = cfg["folder"]
    if cfg["scope"] == "clinic" and clinic and clinic != "all":
        return f"{base}/{clinic}/"
    return f"{base}/"


# =============================================================================
# Clinic-view filtering — applied by the report endpoint AFTER parsing, BEFORE
# anything is sent to the browser. A clinic user's browser never receives other
# clinics' rows or hidden tabs.
# =============================================================================

def _bcba_tab_matches(tab_name: str, clinic_set: set) -> bool:
    """True if this BCBA tab belongs to one of the user's clinics.
    Accepts the full clinic name or a legacy short code."""
    n = str(tab_name).strip()
    if _norm(n) in clinic_set:
        return True
    full = BCBA_TAB_CODE_TO_CLINIC.get(n.upper())
    return full is not None and _norm(full) in clinic_set


def _filter_sheet_rows(rows: list, column: str, clinic_set: set):
    """
    Keep title/header rows up to and including the header row (the first row
    containing `column` exactly), then only data rows whose value in that
    column matches one of the user's clinics. Returns None if the header
    column can't be found (fail closed: caller drops the sheet).
    """
    header_idx = None
    col_idx = None
    for i, row in enumerate(rows):
        for j, cell in enumerate(row):
            if isinstance(cell, str) and cell.strip() == column:
                header_idx, col_idx = i, j
                break
        if header_idx is not None:
            break
    if header_idx is None:
        return None

    kept = list(rows[: header_idx + 1])
    for row in rows[header_idx + 1:]:
        val = row[col_idx] if col_idx < len(row) else None
        if val is not None and _norm(val) in clinic_set:
            kept.append(row)
    return kept


def apply_clinic_view(parsed: dict, report_key: str, user_clinics: list) -> dict:
    """
    Reduce a parsed workbook payload ({sheets, charts, meta}) to what a
    clinic-scoped user may see, per CLINIC_TAB_RULES. Unknown tabs are removed
    (default deny). Charts are kept only for whole ('all') tabs — charts on
    filtered tabs reference the original row layout and would render wrong.
    """
    clinic_set = {_norm(c) for c in user_clinics}
    rules = CLINIC_TAB_RULES.get(report_key)
    out_sheets = []
    whole_tabs = set()

    if rules == "own_tab":
        for sheet in parsed.get("sheets", []):
            if _bcba_tab_matches(sheet.get("name", ""), clinic_set):
                out_sheets.append(sheet)
                whole_tabs.add(sheet.get("name"))
    elif isinstance(rules, dict):
        for sheet in parsed.get("sheets", []):
            rule = rules.get(str(sheet.get("name", "")).strip())
            if not rule:
                continue  # default deny
            if rule[0] == "all":
                out_sheets.append(sheet)
                whole_tabs.add(sheet.get("name"))
            elif rule[0] == "filter":
                new_rows = _filter_sheet_rows(sheet.get("rows", []), rule[1], clinic_set)
                if new_rows is not None:
                    out_sheets.append({"name": sheet.get("name"), "rows": new_rows})
    # rules is None -> report shouldn't be reachable for clinic users; return nothing.

    out_charts = [
        ch for ch in parsed.get("charts", [])
        if ch.get("sheet") in whole_tabs and ch.get("catSheet") in whole_tabs
    ]

    return {"sheets": out_sheets, "charts": out_charts, "meta": parsed.get("meta", {})}


def build_filtered_xlsx(filtered: dict) -> bytes:
    """
    Rebuild an .xlsx from an already-filtered payload (apply_clinic_view
    output) so clinic users can download exactly what they see on screen —
    plain formatting, no charts, no other clinics' data.
    """
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    for sheet in filtered.get("sheets", []):
        ws = wb.create_sheet(title=str(sheet.get("name", "Sheet"))[:31])
        widths = {}
        for r, row in enumerate(sheet.get("rows", []), start=1):
            for c, val in enumerate(row, start=1):
                if val is None:
                    continue
                cell = ws.cell(row=r, column=c, value=val)
                if r <= 4 and isinstance(val, str):
                    cell.font = Font(bold=True)
                widths[c] = max(widths.get(c, 0), len(str(val)))
        for c, w in widths.items():
            ws.column_dimensions[get_column_letter(c)].width = min(max(w + 2, 8), 40)
    if not wb.sheetnames:
        wb.create_sheet(title="No data")
    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
